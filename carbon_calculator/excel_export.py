# -*- coding: utf-8 -*-
"""
Carbon1 (자생복원종 교목/관목) 결과의 Excel(.xlsx) 내보내기.

openpyxl 에만 의존하며 PyQt/matplotlib 에 의존하지 않는다(테스트 용이).
호출측(main_window / combined_window)이 순수 파이썬 dict(`payload`)를 만들어 넘긴다.

단일 지역 export (`export_carbon1_to_excel`) — 4 sheets:
  추정_교목 · 추정_관목 · 기여도 · 그래프

전체 지역 통합 export (`export_all_regions_to_excel`) — 4 sheets:
  지역별_추정치 · 탄소_기여도 · 지역_비교분석 · 그래프

payload 구조 (단일 지역)::

    {
      "tree":  <category_dict>,
      "shrub": <category_dict>,
      "grand_total": float,
      "generated_at": "YYYY-MM-DD HH:MM:SS",
      "region": {"name", "environment", "area_w", "area_h", "area"},
      "images": [{"title", "png": bytes}, ...],
    }

    category_dict = {
      "unit": "cm" | "mm",
      "total": float,
      "total_qty": int,
      "rows": [ {species, diameter, quantity, carbon, checked,
                 a, b, cf, dmin, dmax}, ... ],
      "projection": None | {
          "years": [int, ...],
          "series": [ {label, checked, carbon: [float, ...]}, ... ],
          "total": None | [float, ...],
      },
      "contribution": [ {species, carbon, percent}, ... ],
    }
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .i18n import environment_name, species_name, tr

try:
    from openpyxl.drawing.image import Image as XLImage  # Pillow 필요
except Exception:  # noqa: BLE001 — Pillow 미설치 시 이미지 임베드만 생략
    XLImage = None


# ─────────────────────────── 스타일 ───────────────────────────
HEADER_FILL     = PatternFill("solid", fgColor="2E8B57")
HEADER_FONT     = Font(bold=True, color="FFFFFF")
SUB_HEADER_FILL = PatternFill("solid", fgColor="52A870")   # 보조 헤더 (연한 초록)
TITLE_FONT      = Font(bold=True, size=13, color="246B43")
BOLD_FONT       = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
_THIN  = Side(style="thin", color="D8DEE4")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _categories():
    """(표시명, 내부키) — 언어 설정 후에 평가되도록 함수로 둔다."""
    return ((tr("교목"), "tree"), (tr("관목"), "shrub"))


# ─────────────────────────── 공통 헬퍼 ───────────────────────────

def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER


def _set_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _num(cell, value, fmt: str = "#,##0.00") -> None:
    cell.value         = value
    cell.number_format = fmt
    cell.alignment     = CENTER
    cell.border        = BORDER


# ─────────────────────────── 추정 시트 ───────────────────────────

def _fill_projection_ws(ws, proj, kind_kr: str) -> None:
    """projection 데이터를 이미 생성된 워크시트에 기입."""
    if not proj or not proj.get("years"):
        ws["A1"] = tr("{kind} 추정 데이터가 없습니다. (항목을 추가하고 계산하세요)").format(kind=kind_kr)
        ws["A1"].font = TITLE_FONT
        _set_widths(ws, [50])
        return

    years      = proj["years"]
    series     = proj["series"]
    total      = proj.get("total")
    show_total = total is not None

    headers = [tr("경과연도(년)")]
    headers += [s["label"] + ("" if s["checked"] else tr(" (미표시)")) for s in series]
    if show_total:
        headers.append(tr("총 탄소저장량(전체 합)"))
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    _style_header_row(ws, 1, len(headers))

    for i, yr in enumerate(years):
        ws.cell(i + 2, 1, int(yr)).alignment = CENTER
        ws.cell(i + 2, 1).border = BORDER
        col = 2
        for s in series:
            val = float(s["carbon"][i]) if i < len(s["carbon"]) else 0.0
            _num(ws.cell(i + 2, col), round(val, 4), "#,##0.0000")
            col += 1
        if show_total:
            val = float(total[i]) if i < len(total) else 0.0
            _num(ws.cell(i + 2, col), round(val, 4), "#,##0.0000")

    _set_widths(ws, [12] + [20] * (len(headers) - 1))


def _sheet_projection(wb: Workbook, payload: dict, key: str, kind_kr: str) -> None:
    ws = wb.create_sheet(tr("추정_{kind}").format(kind=kind_kr))
    _fill_projection_ws(ws, payload[key].get("projection"), kind_kr)


# ─────────────────────────── 기여도 시트 ───────────────────────────

def _sheet_contribution(wb: Workbook, payload: dict) -> None:
    ws = wb.create_sheet(tr("기여도"))
    headers = [tr("구분"), tr("수종"), tr("탄소량(kgC)"), tr("비율(%)")]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    _style_header_row(ws, 1, len(headers))

    r = 2
    any_row = False
    for kind_kr, key in _categories():
        for item in payload[key].get("contribution", []):
            any_row = True
            ws.cell(r, 1, kind_kr).alignment = CENTER
            ws.cell(r, 2, species_name(item["species"])).alignment = LEFT
            _num(ws.cell(r, 3), round(item["carbon"],  2), "#,##0.00")
            _num(ws.cell(r, 4), round(item["percent"], 2), "0.00")
            for c in (1, 2):
                ws.cell(r, c).border = BORDER
            r += 1
    if not any_row:
        ws.cell(2, 1, tr("기여도 데이터가 없습니다."))

    _set_widths(ws, [8, 30, 16, 12])


# ─────────────────────────── 그래프 시트 (단일 지역) ───────────────────────────

def _sheet_graphs(wb: Workbook, payload: dict) -> None:
    """그래프(추정/기여도) PNG 이미지를 한 시트에 임베드."""
    ws = wb.create_sheet(tr("그래프"))
    ws.column_dimensions["A"].width = 18
    images = payload.get("images") or []
    if XLImage is None:
        ws["A1"] = tr("이미지 임베드 불가 (Pillow 미설치)")
        return
    if not images:
        ws["A1"] = tr("그래프 이미지가 없습니다.")
        return

    row = 1
    for item in images:
        ws.cell(row, 1, item.get("title", tr("그래프"))).font = TITLE_FONT
        try:
            img = XLImage(io.BytesIO(item["png"]))
            target_w = 760
            if img.width:
                ratio      = target_w / float(img.width)
                img.width  = target_w
                img.height = int(img.height * ratio)
            ws.add_image(img, f"A{row + 1}")
            span = max(18, int((img.height or 360) / 18) + 3)
        except Exception:  # noqa: BLE001
            ws.cell(row + 1, 1, tr("(이미지 로드 실패)"))
            span = 3
        row += span


# ─────────────────────────── 진입점 (단일 지역) ───────────────────────────

def export_carbon1_to_excel(path: str, payload: dict) -> None:
    """payload 를 4개 시트(.xlsx)로 저장(그래프 이미지 포함). 실패 시 예외를 그대로 전파.

    Sheets: 추정_교목 · 추정_관목 · 기여도 · 그래프
    """
    wb = Workbook()
    # 첫 번째 시트(Workbook 기본 활성 시트)를 교목 추정으로 사용
    ws        = wb.active
    ws.title  = tr("추정_교목")
    _fill_projection_ws(ws, payload["tree"].get("projection"), tr("교목"))
    _sheet_projection(wb, payload, "shrub", tr("관목"))
    _sheet_contribution(wb, payload)
    _sheet_graphs(wb, payload)
    wb.save(path)


# ─────────────────────────── 전체 지역 통합 내보내기 ───────────────────────────

def _category_totals_by_year(cat_payload: dict) -> tuple:
    """카테고리 payload 에서 (years, totals) 반환. totals = 모든 수종 합산."""
    proj = cat_payload.get("projection")
    if not proj or not proj.get("years"):
        return [], []
    years  = proj["years"]
    series = proj.get("series", [])
    if not series:
        return years, [0.0] * len(years)
    n      = len(years)
    totals = [
        sum(float(s["carbon"][i]) if i < len(s["carbon"]) else 0.0 for s in series)
        for i in range(n)
    ]
    return years, totals


def _sheet_all_projections(wb: Workbook, payloads: list) -> None:
    """지역별_추정치: 경과연도 × (교목합계 · 관목합계 · 총합계) per region."""
    ws       = wb.active
    ws.title = tr("지역별_추정치")

    # ── 지역별 데이터 수집 ─────────────────────────────────────────────────
    region_data = []
    ref_years   = None
    for p in payloads:
        name         = (p.get("region") or {}).get("name") or tr("지역")
        ty, tt       = _category_totals_by_year(p["tree"])
        sy, st       = _category_totals_by_year(p["shrub"])
        years        = ty if ty else sy
        if ref_years is None and years:
            ref_years = years
        region_data.append({"name": name, "tree": tt, "shrub": st})

    if ref_years is None:
        ws["A1"] = tr("추정 데이터가 없습니다.")
        ws["A1"].font = TITLE_FONT
        return

    n = len(ref_years)
    for rd in region_data:                              # 길이 통일 (짧으면 0 패딩)
        rd["tree"]  = (rd["tree"]  + [0.0] * n)[:n]
        rd["shrub"] = (rd["shrub"] + [0.0] * n)[:n]

    # ── 헤더 ───────────────────────────────────────────────────────────────
    n_regions  = len(region_data)
    total_cols = 1 + n_regions * 3

    # Row 1: 제목 (전체 열 병합)
    ws.cell(1, 1, tr("지역별 탄소저장량 추정치 (경과연도별)")).font = TITLE_FONT
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Row 2~3: 경과연도(년) 2·3행 병합 + 지역명 3열씩 병합
    c       = ws.cell(2, 1, tr("경과연도(년)"))
    c.fill  = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

    col = 2
    for rd in region_data:
        c      = ws.cell(2, col, rd["name"])
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        col += 3

    # Row 3: 교목합계 · 관목합계 · 총합계 (지역별 반복)
    col = 2
    for _ in region_data:
        for sub in [tr("교목합계(kgC)"), tr("관목합계(kgC)"), tr("총합계(kgC)")]:
            c      = ws.cell(3, col, sub)
            c.fill = SUB_HEADER_FILL
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = CENTER; c.border = BORDER
            col += 1

    # ── 데이터 행 ──────────────────────────────────────────────────────────
    for i, yr in enumerate(ref_years):
        row       = i + 4
        c         = ws.cell(row, 1, int(yr))
        c.alignment = CENTER; c.border = BORDER
        col = 2
        for rd in region_data:
            tv = rd["tree"][i]
            sv = rd["shrub"][i]
            _num(ws.cell(row, col),     round(tv,      2)); col += 1
            _num(ws.cell(row, col),     round(sv,      2)); col += 1
            _num(ws.cell(row, col),     round(tv + sv, 2)); col += 1

    _set_widths(ws, [12] + [18, 18, 18] * n_regions)
    ws.freeze_panes = ws.cell(4, 2)


def _sheet_all_contributions(wb: Workbook, payloads: list) -> None:
    """탄소_기여도: 모든 지역의 수종별 기여도를 지역 구분 헤더를 붙여 나열."""
    ws = wb.create_sheet(tr("탄소_기여도"))
    ws.cell(1, 1, tr("지역별 수종 탄소 기여도")).font = TITLE_FONT

    r = 3
    for p in payloads:
        name = (p.get("region") or {}).get("name") or tr("지역")

        # 지역 구분 헤더
        rc           = ws.cell(r, 1, f"▶  {name}")
        rc.font      = Font(bold=True, size=12, color="246B43")
        rc.alignment = LEFT
        r += 1

        # 열 헤더
        for c, h in enumerate([tr("구분"), tr("수종"), tr("탄소량(kgC)"), tr("비율(%)")], 1):
            ws.cell(r, c, h)
        _style_header_row(ws, r, 4)
        r += 1

        any_row = False
        for kind_kr, key in _categories():
            for item in p[key].get("contribution", []):
                any_row = True
                ws.cell(r, 1, kind_kr).alignment = CENTER
                ws.cell(r, 2, species_name(item["species"])).alignment = LEFT
                _num(ws.cell(r, 3), round(item["carbon"],  2), "#,##0.00")
                _num(ws.cell(r, 4), round(item["percent"], 2), "0.00")
                for ci in (1, 2):
                    ws.cell(r, ci).border = BORDER
                r += 1

        if not any_row:
            ws.cell(r, 1, tr("기여도 데이터가 없습니다.")).alignment = LEFT
            r += 1

        r += 1  # 빈 줄 구분

    _set_widths(ws, [8, 30, 16, 12])


def _sheet_comparison(wb: Workbook, comparison_data: list) -> None:
    """지역_비교분석: 지역별 총 탄소저장량 비교 테이블."""
    ws = wb.create_sheet(tr("지역_비교분석"))

    ws.cell(1, 1, tr("지역별 총 탄소저장량 비교 분석")).font = TITLE_FONT
    ws.merge_cells("A1:G1")

    headers = [tr("지역명"), tr("면적(㎡)"), tr("환경"), tr("교목(kgC)"), tr("관목(kgC)"),
               tr("총 탄소저장량(kgC)"), tr("단위면적당(kgC/㎡)")]
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)
    _style_header_row(ws, 2, len(headers))

    for i, d in enumerate(comparison_data, start=3):
        ws.cell(i, 1, d["name"]).alignment = LEFT
        ws.cell(i, 1).border = BORDER
        _num(ws.cell(i, 2), d["area"],             "#,##0")
        ws.cell(i, 3, environment_name(d["env"])).alignment = LEFT
        ws.cell(i, 3).border = BORDER
        _num(ws.cell(i, 4), round(d["tree"],    2))
        _num(ws.cell(i, 5), round(d["shrub"],   2))
        _num(ws.cell(i, 6), round(d["total"],   2))
        _num(ws.cell(i, 7), round(d["density"], 4), "#,##0.0000")

    if comparison_data:
        sr      = len(comparison_data) + 3
        t_tree  = sum(d["tree"]  for d in comparison_data)
        t_shrub = sum(d["shrub"] for d in comparison_data)
        grand   = t_tree + t_shrub
        c       = ws.cell(sr, 1, tr("전체 합계"))
        c.font  = BOLD_FONT; c.alignment = LEFT; c.border = BORDER
        _num(ws.cell(sr, 4), round(t_tree,  2))
        _num(ws.cell(sr, 5), round(t_shrub, 2))
        _num(ws.cell(sr, 6), round(grand,   2))
        for ci in (2, 3, 7):
            ws.cell(sr, ci).border = BORDER

    _set_widths(ws, [16, 12, 24, 16, 16, 20, 18])


def _render_comparison_chart_png(comparison_data: list):
    """지역별 탄소저장량 비교 막대 차트를 PNG 바이트로 렌더링 (Agg 백엔드, Qt 불필요)."""
    if not comparison_data:
        return None
    try:
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        from matplotlib.figure import Figure
        import matplotlib.pyplot as plt

        try:
            plt.rcParams["font.family"]       = "Malgun Gothic"
            plt.rcParams["axes.unicode_minus"] = False
        except Exception:
            pass

        names      = [d["name"]  for d in comparison_data]
        tree_vals  = [d["tree"]  for d in comparison_data]
        shrub_vals = [d["shrub"] for d in comparison_data]
        totals     = [t + s for t, s in zip(tree_vals, shrub_vals)]

        fig    = Figure(figsize=(12, 5), layout="constrained")
        canvas = FigureCanvasAgg(fig)
        ax     = fig.add_subplot(111)

        cmap    = plt.get_cmap("tab10")
        hatches = ["//", "\\\\", "..", "xx", "++", "oo", "--", "**", "||"]

        for i, (tot, name) in enumerate(zip(totals, names)):
            ax.bar(i, tot, width=0.7,
                   color=cmap(i % 10), hatch=hatches[i % len(hatches)],
                   edgecolor="white", linewidth=0.8, label=name)

        top = max(totals) if totals else 1.0
        for i, tot in enumerate(totals):
            ax.text(i, tot + (top * 0.01 if top else 0.0), f"{tot:,.1f}",
                    ha="center", va="bottom", fontsize=11, fontweight="bold")

        ax.set_xticks([])
        ax.set_ylabel(tr("탄소저장량 (kgC)"))
        ax.set_ylim(0, top * 1.15 if top > 0 else 1.0)
        ax.set_title(tr("지역별 총 탄소저장량 비교"), fontweight="bold")
        ncol = 1 + (len(names) - 1) // 12
        ax.legend(title=tr("지역"), loc="upper left", bbox_to_anchor=(1.01, 1.0),
                  fontsize=10, framealpha=0.95, borderaxespad=0.0, ncol=max(1, ncol))
        ax.grid(True, axis="y", alpha=0.3)

        buf = io.BytesIO()
        canvas.print_png(buf)
        return buf.getvalue()
    except Exception:  # noqa: BLE001
        return None


def _sheet_all_graphs(wb: Workbook, payloads: list, comparison_data: list) -> None:
    """그래프: 지역 비교 차트를 맨 위에, 이어서 지역별 그래프를 임베드."""
    ws = wb.create_sheet(tr("그래프"))
    ws.column_dimensions["A"].width = 18

    if XLImage is None:
        ws["A1"] = tr("이미지 임베드 불가 (Pillow 미설치)")
        return

    row = 1

    # 1) 지역 비교 막대 차트
    cmp_png = _render_comparison_chart_png(comparison_data)
    if cmp_png:
        ws.cell(row, 1, tr("지역별 탄소저장량 비교")).font = TITLE_FONT
        try:
            img = XLImage(io.BytesIO(cmp_png))
            target_w   = 900
            if img.width:
                ratio      = target_w / float(img.width)
                img.width  = target_w
                img.height = int(img.height * ratio)
            ws.add_image(img, f"A{row + 1}")
            span = max(20, int((img.height or 360) / 18) + 3)
        except Exception:  # noqa: BLE001
            ws.cell(row + 1, 1, tr("(이미지 로드 실패)"))
            span = 3
        row += span

    # 2) 지역별 그래프
    for p in payloads:
        region_name = (p.get("region") or {}).get("name") or tr("지역")
        images      = p.get("images") or []
        if not images:
            continue

        c      = ws.cell(row, 1, tr("── {name} 지역 ──").format(name=region_name))
        c.font = Font(bold=True, size=12, color="246B43")
        row   += 1

        for item in images:
            ws.cell(row, 1, item.get("title", tr("그래프"))).font = TITLE_FONT
            try:
                img = XLImage(io.BytesIO(item["png"]))
                target_w   = 760
                if img.width:
                    ratio      = target_w / float(img.width)
                    img.width  = target_w
                    img.height = int(img.height * ratio)
                ws.add_image(img, f"A{row + 1}")
                span = max(18, int((img.height or 360) / 18) + 3)
            except Exception:  # noqa: BLE001
                ws.cell(row + 1, 1, tr("(이미지 로드 실패)"))
                span = 3
            row += span


# ─────────────────────────── 진입점 (전체 지역 통합) ───────────────────────────

def export_all_regions_to_excel(path: str, payloads: list, comparison_data: list) -> None:
    """모든 지역 통합 결과를 4개 시트(.xlsx)로 저장. 실패 시 예외를 그대로 전파.

    Sheets: 지역별_추정치 · 탄소_기여도 · 지역_비교분석 · 그래프

    Args:
        payloads:        각 지역의 export payload 리스트 (region, tree, shrub, images 포함).
        comparison_data: 지역 비교 데이터 리스트
                         [{name, area, env, tree, shrub, total, density}, ...].
    """
    wb = Workbook()
    _sheet_all_projections(wb, payloads)
    _sheet_all_contributions(wb, payloads)
    _sheet_comparison(wb, comparison_data)
    _sheet_all_graphs(wb, payloads, comparison_data)
    wb.save(path)
